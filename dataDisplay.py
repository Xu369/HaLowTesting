import re
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference


def extract_info(filename):
    with open(filename, "r") as file:
        data = file.read()

    # Regular expression: channel, bitrate, sender, receiver
    pattern = re.compile(
        r"channel (\d+)\nbitrate (\S+)\n(\d+)/(\d+) \(\d+%\) sender\n(\d+)/(\d+) \(\d+(\.\d+)?+%\) receiver",
        re.IGNORECASE,
    )

    results = []

    matches = pattern.findall(data)
    for match in matches:
        channel = int(match[0])
        bitrate = match[1]
        sender_lost = int(match[2])
        sender_total = int(match[3])
        receiver_lost = int(match[4])
        receiver_total = int(match[5])

        results.append(
            {
                "channel": channel,
                "bitrate": bitrate,
                "sender_lost": sender_lost,
                "sender_total": sender_total,
                "receiver_lost": receiver_lost,
                "receiver_total": receiver_total,
            }
        )

    return results


def export_to_excel(data, output_file):
    df = pd.DataFrame(data)

    # Calculate packet loss rate
    df["sender_loss_rate"] = df["sender_lost"] / df["sender_total"] * 100
    df["receiver_loss_rate"] = df["receiver_lost"] / df["receiver_total"] * 100

    # Uniform units (kbps)
    df["bitrate_kbps"] = df["bitrate"].apply(
        lambda x: (
            int(x[:-1]) * 1000
            if x[-1].lower() == "m"
            else (int(x[:-1]) if x[-1].lower() == "k" else int(x[:-1]) * 1000000)
        )
    )

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Packet Loss Data"

    # Append data to Excel
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # 创建图表
    chart = LineChart()
    chart.title = "Packet Loss Rate by Channel and Bitrate"
    chart.style = 13
    chart.x_axis.title = "Bitrate (kbps)"
    chart.y_axis.title = "Loss Rate (%)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100

    # Add data to chart
    data = Reference(ws, min_col=8, min_row=1, max_col=9, max_row=ws.max_row)
    categories = Reference(ws, min_col=7, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Add chart to Excel workbook
    ws.add_chart(chart, "L10")

    wb.save(output_file)
    print(f"Data and chart have been written to {output_file}")


if __name__ == "__main__":
    filename = "output_Data/security_frequency_bandwidth.txt"
    output_file = "excelDisplay/security_frequency_bandwidth.xlsx"
    data = extract_info(filename)
    export_to_excel(data, output_file)
